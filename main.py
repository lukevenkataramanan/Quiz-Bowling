import pygame
import openpyxl
from openpyxl import utils
import random
pygame.init()
pygame.font.init()

pygame.display.set_caption('Quiz Bowling')
WINDOW_WIDTH = 1440
WINDOW_HEIGHT = 780
win = pygame.display.set_mode((WINDOW_WIDTH, WINDOW_HEIGHT))

BUTTON_WIDTH = 400
BUTTON_HEIGHT = 75
active_window = 0
# -1 not running
# 0 start screen
# 1 leaderboard
# 2 game

game_speed = 10
score = 0
player_name = ''
SPAWN_OFFSET = 10  # arbitrary number to get around pin spawning bug
FONT = pygame.font.SysFont("Trebuchet MS", 65)
SCORE_FONT = pygame.font.SysFont("copperplate", 50)

SHIP_IMG = pygame.transform.scale(pygame.image.load('images/avatar.png'), (100, 100))
BASE_IMG = pygame.transform.scale(pygame.image.load('images/bowling bumpers.png'), (1440, 75))
BG_IMG = pygame.transform.scale(pygame.image.load('images/bowling lane.png'), (1440, 800))
OBS_IMG = pygame.transform.scale(pygame.image.load('images/bowling pinsetter.png'), (100, 640))
PIN_IMG = pygame.transform.scale(pygame.image.load('images/bowling pin.png'), (100, 200))
ball_dim = 75
RED_BALL_IMG = pygame.transform.scale(pygame.image.load('images/red bowling ball.png'), (ball_dim, ball_dim))
GREEN_BALL_IMG = pygame.transform.scale(pygame.image.load('images/green bowling ball.png'), (ball_dim, ball_dim))
BLUE_BALL_IMG = pygame.transform.scale(pygame.image.load('images/blue bowling ball.png'), (ball_dim, ball_dim))
PURPLE_BALL_IMG = pygame.transform.scale(pygame.image.load('images/purple bowling ball.png'), (ball_dim, ball_dim))
button_dim = 100
RED_BUTTON_IMG = pygame.transform.scale(pygame.image.load('images/red button.png'), (button_dim, button_dim))
GREEN_BUTTON_IMG = pygame.transform.scale(pygame.image.load('images/green button.png'), (button_dim, button_dim))
BLUE_BUTTON_IMG = pygame.transform.scale(pygame.image.load('images/blue button.png'), (button_dim, button_dim))
PURPLE_BUTTON_IMG = pygame.transform.scale(pygame.image.load('images/purple button.png'), (button_dim, button_dim))

colors = {
    'menu dark purple': (80, 41, 137),
    'purple': (159, 4, 197),
    'dark gray': (54, 58, 55),
    'html gray': (128, 128, 128),
    'html dark gray': (169, 169, 169),
    'black': (0, 0, 0),
    'white': (255, 255, 255)
}


def start_screen():
    global active_window, player_name
    # WINDOW_WIDTH = 1000
    # WINDOW_HEIGHT = 800
    intro = True
    username = ''
    player_name = ''
    active_text_field = True
    click = False
    title_font = pygame.font.SysFont('Segoe UI', 100, True)  # create font
    secondary_font = pygame.font.SysFont('sfprotextthin', 65)  # create font
    tertiary_font = pygame.font.SysFont('trebuchetmsitalic', 55)  # create font

    while intro:
        pygame.draw.rect(win, colors['menu dark purple'], (0, 0, 1440, 800))  # fill in color

        title_text = title_font.render('Quiz Bowling', True, colors['white'])  # create text
        win.blit(title_text, (WINDOW_WIDTH/2 - title_text.get_width() / 2, 30))  # draw title text

        win.blit(RED_BALL_IMG, (WINDOW_WIDTH / 2 - PURPLE_BALL_IMG.get_width()*2.15 - 15, 300))
        win.blit(GREEN_BALL_IMG, (WINDOW_WIDTH / 2 - PURPLE_BALL_IMG.get_width()*1.15 - 5, 300))
        win.blit(BLUE_BALL_IMG, (WINDOW_WIDTH / 2 - PURPLE_BALL_IMG.get_width()*0.15 + 5, 300))
        win.blit(PURPLE_BALL_IMG, (WINDOW_WIDTH / 2 + PURPLE_BALL_IMG.get_width()*0.85 + 15, 300))

        mx, my = pygame.mouse.get_pos()  # get mouse position

        start_button = pygame.Rect(WINDOW_WIDTH/2 - BUTTON_WIDTH/2, 400, BUTTON_WIDTH, BUTTON_HEIGHT)
        start_button_color = colors['black']
        start_text = secondary_font.render('START', True, colors['white'])
        leaderboard_button = pygame.Rect(WINDOW_WIDTH / 2 - BUTTON_WIDTH / 2, 500, BUTTON_WIDTH, BUTTON_HEIGHT)
        leaderboard_button_color = colors['black']
        leaderboard_text = secondary_font.render('LEADERBOARD', True, colors['white'])
        exit_button = pygame.Rect(WINDOW_WIDTH/2 - BUTTON_WIDTH/2, 600, BUTTON_WIDTH, BUTTON_HEIGHT)
        exit_button_color = colors['black']
        exit_text = secondary_font.render('EXIT', True, colors['white'])
        if start_button.collidepoint((mx, my)):
            start_button_color = colors['dark gray']
            if click and username != '':
                player_name = username
                active_window = 2
                intro = False
        if leaderboard_button.collidepoint((mx, my)):
            leaderboard_button_color = colors['dark gray']
            if click:  # if click on start button, show leaderboards
                active_window = 1
                intro = False
        if exit_button.collidepoint((mx, my)):
            exit_button_color = colors['dark gray']
            if click:  # if click on stop button
                active_window = -1
                intro = False

        # Start button
        pygame.draw.rect(win, colors['white'],
                         (WINDOW_WIDTH / 2 - (BUTTON_WIDTH + 10) / 2, 395, BUTTON_WIDTH + 10,
                          BUTTON_HEIGHT + 10))
        pygame.draw.rect(win, start_button_color, start_button)  # draw start button
        win.blit(start_text,
                 (WINDOW_WIDTH / 2 - start_text.get_width() / 2,
                  start_button.y + start_button.height / 2 - start_text.get_height() / 2))
        # Leaderboard button
        pygame.draw.rect(win, colors['white'],
                         (WINDOW_WIDTH / 2 - (BUTTON_WIDTH + 10) / 2, 495, BUTTON_WIDTH + 10, BUTTON_HEIGHT + 10))
        pygame.draw.rect(win, leaderboard_button_color, leaderboard_button)  # draw leaderboard button
        win.blit(leaderboard_text,
                 (WINDOW_WIDTH / 2 - leaderboard_text.get_width() / 2,
                  leaderboard_button.y + leaderboard_button.height / 2 - leaderboard_text.get_height() / 2))
        # Exit button
        pygame.draw.rect(win, colors['white'],
                         (WINDOW_WIDTH/2 - (BUTTON_WIDTH+10)/2, 595, BUTTON_WIDTH+10, BUTTON_HEIGHT+10))
        pygame.draw.rect(win, exit_button_color, exit_button)  # draw exit button
        win.blit(exit_text,
                 (WINDOW_WIDTH/2 - exit_text.get_width()/2,
                  exit_button.y + exit_button.height/2-exit_text.get_height()/2))

        # Username
        pygame.draw.rect(win, colors['purple'],
                         (WINDOW_WIDTH/2 - (BUTTON_WIDTH+10)/2, 195, BUTTON_WIDTH+10, BUTTON_HEIGHT+10))
        username_textfield = pygame.Rect(WINDOW_WIDTH/2 - BUTTON_WIDTH/2, 200, BUTTON_WIDTH, BUTTON_HEIGHT)
        username_textfield_color = colors['html dark gray'] if active_text_field else colors['html gray']
        pygame.draw.rect(win, username_textfield_color, username_textfield)
        username_text_color = colors['black'] if username != '' else colors['white']
        username_text = tertiary_font.render(username, True, username_text_color)
        if username == '':
            username_text = tertiary_font.render('Enter name', True, username_text_color)
        win.blit(username_text,
                 (WINDOW_WIDTH / 2 - username_text.get_width() / 2,
                  username_textfield.y + username_textfield.height / 2 - username_text.get_height() / 2))

        credit_text = tertiary_font.render('Created by Luke Venkataramanan®', True, colors['white'])
        win.blit(credit_text, credit_text.get_rect(center=(720, 750)))

        if click:
            active_text_field = False
            if username_textfield.collidepoint((mx, my)):
                active_text_field = True

        click = False
        for event in pygame.event.get():
            if event.type == pygame.QUIT:  # if quit, don't start game and then quit
                active_window = -1
                intro = False
            elif event.type == pygame.MOUSEBUTTONDOWN:  # if click down, set click to true
                if event.button == 1:
                    click = True
            elif active_text_field and event.type == pygame.KEYDOWN:
                if event.key == pygame.K_BACKSPACE:  # hit backspace
                    username = username[:-1]
                elif username != '' and event.key == pygame.K_RETURN:
                    active_window = 2
                    player_name = username
                    intro = False
                else:  # type in a character
                    username += event.unicode
        pygame.display.update()
    return active_window


def leaderboard():
    global active_window
    hs = get_high_scores()
    lb = True
    click = False
    title_font = pygame.font.SysFont('Segoe UI', 100, True)  # create font
    secondary_font = pygame.font.SysFont('sfprotextthin', 65)  # create font
    tertiary_font = pygame.font.SysFont('sfprotextthin', 40)  # create font

    while lb:
        pygame.draw.rect(win, colors['menu dark purple'], (0, 0, 1440, 800))  # fill in color

        title_text = title_font.render('Leaderboard', True, colors['white'])  # create text
        win.blit(title_text, (WINDOW_WIDTH / 2 - title_text.get_width() / 2, 30))  # draw title text

        mx, my = pygame.mouse.get_pos()  # get mouse position

        return_button = pygame.Rect(WINDOW_WIDTH / 2 - BUTTON_WIDTH / 2, 600, BUTTON_WIDTH, BUTTON_HEIGHT)
        return_button_color = colors['black']
        return_text = secondary_font.render('RETURN', True, colors['white'])
        if return_button.collidepoint((mx, my)):
            return_button_color = colors['dark gray']
            if click:  # if click on return, return to start screen
                active_window = 0
                lb = False

        # Return button
        pygame.draw.rect(win, colors['white'],
                         (WINDOW_WIDTH / 2 - (BUTTON_WIDTH + 10) / 2, 595, BUTTON_WIDTH + 10, BUTTON_HEIGHT + 10))
        pygame.draw.rect(win, return_button_color, return_button)  # draw return button
        win.blit(return_text,
                 (WINDOW_WIDTH / 2 - return_text.get_width() / 2,
                  return_button.y + return_button.height / 2 - return_text.get_height() / 2))
        click = False
        for event in pygame.event.get():
            if event.type == pygame.QUIT:  # if quit, don't start game and then quit
                active_window = -1
                lb = False
            elif event.type == pygame.MOUSEBUTTONDOWN:  # if click down, set click to true
                if event.button == 1:
                    click = True

        # Display Scores
        win.blit(tertiary_font.render('Rank:', True, colors['white']), (340, 160))
        win.blit(tertiary_font.render('Player:', True, colors['white']), (460, 160))
        win.blit(tertiary_font.render('Score:', True, colors['white']), (995, 160))
        for i, s in enumerate(hs):
            if i > 4:
                break
            rank_text = secondary_font.render(f'#{str(i + 1)}.   ', True, colors['white'])
            name_text = secondary_font.render(s[1], True, colors['white'])
            score_text = secondary_font.render(str(s[0]), True, colors['white'])
            win.blit(rank_text, (360, i*75+215))
            win.blit(name_text, (360 + rank_text.get_width(), i * 75 + 215))
            win.blit(score_text, (1080-score_text.get_width(), i * 75 + 215))

        pygame.display.update()
    return active_window


def get_high_scores():
    hs = []
    lines_to_remove = []
    with open('highscores.txt', 'r') as file:
        lines = [line.rstrip() for line in file.readlines()]
        for line in lines:
            if line:
                s = line.split(': ')  # separate username and high score
                s[0] = int(s[0])
                if len(s) >= 2:  # must have score and username
                    hs.append(s)
                elif len(s) >= 1:  # if only have score, delete line
                    lines_to_remove.append(line)
    with open('highscores.txt', 'w') as file:  # cleans up incomplete entries
        for line in lines:
            if line not in lines_to_remove:
                file.write(line + '\n')

    hs.sort(reverse=True)  # put highscores into descending order
    return hs


def get_questions():
    """
    parameters for Question(): question (str), answers (List: ints), correct_answer (int)
    :return:
    """
    workbook = openpyxl.load_workbook('QuizBowlingData.xlsx')
    sheet = workbook['Sheet1']
    questions = []
    for row in range(2, sheet.max_row + 1):
        q = sheet['A' + str(row)].value
        answer_choices = []
        for col in range(2, 6):  # cols 2 through 5
            cell = sheet[utils.get_column_letter(col) + str(row)]
            answer_choices.append(cell.value)
        correct = sheet['F' + str(row)].value
        question = Question(q, answer_choices, correct)
        question.shuffle_answers()
        questions.append(question)
    return questions


class Ship:
    MAX_ROTATION = 20
    ROT_VEL = 2
    ACCELERATION = 2
    TERMINAL_VELOCITY = 20
    FRICTION = 1

    def __init__(self, x, y):
        self.x = x
        self.y = y
        self.rot = 0
        self.vel = 0
        self.img = SHIP_IMG

        self.can_shoot = True
        self.ball = None

    def move(self):
        self.y += self.vel

        # max velocity
        if self.vel < 0:  # if ship is moving upwards
            if self.vel < -self.TERMINAL_VELOCITY:
                self.vel = -self.TERMINAL_VELOCITY
            self.vel += self.FRICTION
        elif self.vel > 0:  # if ship is moving downwards
            if self.vel > self.TERMINAL_VELOCITY:
                self.vel = self.TERMINAL_VELOCITY
            self.vel -= self.FRICTION
        # return ship to normal tilt
        if self.rot < 0:
            self.rot += self.FRICTION
        if self.rot > 0:
            self.rot -= self.FRICTION

    def fly(self, direction):
        if direction == 'up':
            self.vel -= self.ACCELERATION
            if self.rot < self.MAX_ROTATION:  # rotate if not fully rotated
                self.rot += self.ROT_VEL

        elif direction == 'down':
            self.vel += self.ACCELERATION
            if self.rot > -self.MAX_ROTATION:  # rotate if not fully rotated
                self.rot -= self.ROT_VEL

    def draw(self):
        rotated_img = pygame.transform.rotate(self.img, self.rot)
        new_rect = rotated_img.get_rect(center=self.img.get_rect(topleft=(self.x, self.y)).center)
        win.blit(rotated_img, new_rect.topleft)

    def get_mask(self):
        return pygame.mask.from_surface(self.img)

    def shoot(self, color):
        self.ball = Ball(self.x, self.y+self.img.get_height()/2-BLUE_BALL_IMG.get_height()/2, color, None)


class Ball:
    """
    spawns in whenever ship calls its shoot method
    """
    VEL = 20
    ROT_VEL = -25

    def __init__(self, x, y, color, text):
        self.x = x  # will be constant but comes from the ship
        self.y = y
        self.rot = 0
        self.color = color
        self.text = text
        if color == 'red':
            self.img = RED_BALL_IMG
        elif color == 'green':
            self.img = GREEN_BALL_IMG
        elif color == 'blue':
            self.img = BLUE_BALL_IMG
        elif color == 'purple':
            self.img = PURPLE_BALL_IMG

    def move(self):
        self.x += self.VEL
        self.rot += self.ROT_VEL

    def draw(self):
        rotated_img = pygame.transform.rotate(self.img, self.rot)
        new_rect = rotated_img.get_rect(center=self.img.get_rect(topleft=(self.x, self.y)).center)
        win.blit(rotated_img, new_rect.topleft)

    def get_mask(self):
        return pygame.mask.from_surface(self.img)


class Obstacle:
    GAP = PIN_IMG.get_height()
    VEL = game_speed

    def __init__(self):
        self.x = 1650

        self.OBS_TOP = pygame.transform.flip(OBS_IMG, False, True)
        self.OBS_BOTTOM = OBS_IMG

        self.height = random.randrange(50, 450)
        self.top = self.height - self.OBS_TOP.get_height()  # y pos of top of upper pipe
        self.bottom = self.height + self.GAP  # y pos of top of bottom pipe

        self.passed = False  # has player gone through pipe yet
        self.pin = Pin(self.x, self.height)  # make pin an attribute of each obstacle

    def move(self):
        self.x -= self.VEL

    def draw(self):
        win.blit(self.OBS_TOP, (self.x, self.top))
        win.blit(self.OBS_BOTTOM, (self.x, self.bottom))

    def collide_with_ship(self, ship):
        ship_mask = ship.get_mask()
        top_mask = pygame.mask.from_surface(self.OBS_TOP)
        bottom_mask = pygame.mask.from_surface(self.OBS_BOTTOM)

        top_offset = ((self.x - ship.x), self.top - round(ship.y))
        bottom_offset = ((self.x - ship.x), self.bottom - round(ship.y))

        top_overlap = ship_mask.overlap(top_mask, top_offset)  # first point of overlap between ship and top obs
        bottom_overlap = ship_mask.overlap(bottom_mask, bottom_offset)  # returns None if not overlapping

        return top_overlap or bottom_overlap

    def collide_with_ball(self, ship):
        ball_mask = pygame.mask.from_surface(ship.ball.img)
        top_mask = pygame.mask.from_surface(self.OBS_TOP)
        bottom_mask = pygame.mask.from_surface(self.OBS_BOTTOM)

        top_offset = ((self.x - ship.ball.x), self.top - round(ship.ball.y))
        bottom_offset = ((self.x - ship.ball.x), self.bottom - round(ship.ball.y))

        top_overlap = ball_mask.overlap(top_mask, top_offset)  # first point of overlap between ship and top obs
        bottom_overlap = ball_mask.overlap(bottom_mask, bottom_offset)  # returns None if not overlapping

        return top_overlap or bottom_overlap


class Pin:
    VEL = game_speed

    def __init__(self, x, y):
        self.x = x + SPAWN_OFFSET  # arbitrary number to get around spawning bug
        self.y = y  # y depends on obstacle it spawns within
        self.img = PIN_IMG

        self.text = ""

        self.destroyed = False

    def move(self):
        self.x -= self.VEL

    def draw(self):
        win.blit(self.img, (self.x, self.y))
        rendered_text = FONT.render(self.text, True, (0, 0, 0))
        text_x = self.x + self.img.get_width()/2 - rendered_text.get_width()/2
        text_y = self.y + self.img.get_height()/2 - rendered_text.get_height()/2
        win.blit(rendered_text, (text_x, text_y))

    def collide_with_ship(self, ship):
        ship_mask = ship.get_mask()
        pin_mask = pygame.mask.from_surface(self.img)

        offset = ((self.x - ship.x), self.y - round(ship.y))

        return ship_mask.overlap(pin_mask, offset)

    def collide_with_ball(self, ship):
        ball_mask = pygame.mask.from_surface(ship.ball.img)
        pin_mask = pygame.mask.from_surface(self.img)

        offset = ((self.x - ship.ball.x), self.y - round(ship.ball.y))

        return ball_mask.overlap(pin_mask, offset)


class Base:
    VEL = game_speed
    WIDTH = BASE_IMG.get_width()
    HEIGHT = BASE_IMG.get_height()

    def __init__(self, top):
        self.img = BASE_IMG if top else pygame.transform.flip(BASE_IMG, False, True)
        self.y = 0 if top else WINDOW_HEIGHT - self.HEIGHT
        self.x1 = 0
        self.x2 = self.WIDTH

    def move(self):
        self.x1 -= self.VEL
        self.x2 -= self.VEL

        if self.x1 + self.WIDTH < 0:
            self.x1 = self.x2 + self.WIDTH

        if self.x2 + self.WIDTH < 0:
            self.x2 = self.x1 + self.WIDTH

    def draw(self):
        win.blit(self.img, (self.x1, self.y))
        win.blit(self.img, (self.x2, self.y))

    def collide_with_ship(self, ship):
        ship_mask = ship.get_mask()
        mask = pygame.mask.from_surface(self.img)

        offset = (0, self.y - round(ship.y))

        return ship_mask.overlap(mask, offset)


class Background:
    WIDTH = BG_IMG.get_width()

    def __init__(self, vel, img):
        self.img = img
        self.vel = vel
        self.y = 0
        self.x1 = 0
        self.x2 = self.WIDTH

    def move(self):
        self.x1 -= self.vel
        self.x2 -= self.vel

        if self.x1 + self.WIDTH < 0:
            self.x1 = self.x2 + self.WIDTH

        if self.x2 + self.WIDTH < 0:
            self.x2 = self.x1 + self.WIDTH

    def draw(self):
        win.blit(self.img, (self.x1, self.y))
        win.blit(self.img, (self.x2, self.y))


class Button:
    def __init__(self, color):
        self.text = ''
        self.x = 10
        y_intercept = 110
        if color == 'red':
            self.img = RED_BUTTON_IMG
            self.y = y_intercept
        elif color == 'green':
            self.img = GREEN_BUTTON_IMG
            self.y = y_intercept + self.img.get_height()*1.5
        elif color == 'blue':
            self.img = BLUE_BUTTON_IMG
            self.y = y_intercept + self.img.get_height()*3
        elif color == 'purple':
            self.img = PURPLE_BUTTON_IMG
            self.y = y_intercept + self.img.get_height()*4.5

    def draw(self):
        win.blit(self.img, (self.x, self.y))
        rendered_text = FONT.render(self.text, True, (255, 255, 255))
        text_x = self.x + self.img.get_width() / 2 - rendered_text.get_width() / 2
        text_y = self.y + self.img.get_height() / 2 - rendered_text.get_height() / 2
        win.blit(rendered_text, (text_x, text_y))


class Question:
    def __init__(self, question, answers, correct_answer):
        """
        :param question: math question
        :param answers: list of 4 answers choices (21, 16, 19, 24)
        :param correct_answer (int): 19
        """
        self.question = question
        self.answers = answers
        self.correct_answer = correct_answer
        self.index_of_correct = answers.index(correct_answer)

    def reset_question(self, pin, buttons):
        pin.text = self.question
        for i, button in enumerate(buttons):  # runs 4 times
            button.text = str(self.answers[i])

    def shuffle_answers(self):
        random.shuffle(self.answers)  # put answer choices in a random order
        self.index_of_correct = self.answers.index(self.correct_answer)


def process_shot(question, ball):
    """
    if pin collided with ball, returns True if the answer choice was the correct answer to the question
    :param ball:
    :param question:
    :return: True if shot by the correct ball, else False
    """
    correct = question.answers.index(question.correct_answer)  # index of correct answer in array of questions
    color = ball.color
    answer_choice_dict = {
        'red': 0,
        'green': 1,
        'blue': 2,
        'purple': 3
    }
    return answer_choice_dict[color] == correct


def update_game_speed(obs, pins, bases):
    global game_speed
    """
    needs to update vel of ob, pin, bases
    obs and pins speeds are updated every time one is created
    :return: 
    """
    game_speed *= 1.1
    for ob in obs:
        ob.VEL = game_speed
    for pin in pins:
        pin.VEL = game_speed
    for base in bases:
        base.VEL = game_speed


def draw_window(ship, top_base, bottom_base, bgs, obs, pins, ball, buttons):
    for bg in bgs:
        bg.draw()
    top_base.draw()
    bottom_base.draw()
    for ob in obs:
        ob.draw()
    for pin in pins:
        pin.draw()
    for button in buttons:
        button.draw()
    ship.draw()
    if ball:
        ball.draw()
    score_text = SCORE_FONT.render("Score: " + str(score), True, (255, 255, 255))
    score_y = 15
    score_x = WINDOW_WIDTH - score_text.get_width() - score_y
    pygame.draw.rect(win, (0, 0, 0), (score_x-5, score_y-5, score_text.get_width()+10, score_text.get_height()+10))
    pygame.draw.rect(win, (0, 0, 255), (score_x, score_y, score_text.get_width(), score_text.get_height()))
    win.blit(score_text, (score_x, score_y))
    pygame.display.update()


def main():
    global game_speed, score, active_window, player_name
    ship = Ship(200, 100)
    top_base = Base(True)
    bottom_base = Base(False)
    bgs = [Background(1, BG_IMG)]
    # add code from Zombie Code 1 in text.txt here for star layers
    buttons = [Button('red'), Button('green'), Button('blue'), Button('purple')]
    obs = [Obstacle()]
    pins = [obs[0].pin]
    pins[0].x -= SPAWN_OFFSET  # arbitrary number to get around spawning bug

    questions = get_questions()
    question = random.choice(questions)
    question.reset_question(pins[0], buttons)

    score = 0
    game_speed = 10

    clock = pygame.time.Clock()

    run = True
    while run:
        clock.tick(30)

        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                run = False

        # Moving ship up and down:
        keys = pygame.key.get_pressed()
        if keys[pygame.K_UP]:
            ship.fly('up')
        if keys[pygame.K_DOWN]:
            ship.fly('down')
        ship.move()

        if ship.can_shoot:
            if keys[pygame.K_1]:
                ship.shoot('red')
                ship.can_shoot = False
            if keys[pygame.K_2]:
                ship.shoot('green')
                ship.can_shoot = False
            if keys[pygame.K_3]:
                ship.shoot('blue')
                ship.can_shoot = False
            if keys[pygame.K_4]:
                ship.shoot('purple')
                ship.can_shoot = False
        if ship.ball:
            ship.ball.move()

        # Obstacles:
        obs_to_remove = []
        add_obs = False
        for ob in obs:
            if ob.collide_with_ship(ship):  # if player hits obstacle
                run = False
            if ship.ball and ob.collide_with_ball(ship):  # if ball hits wall, remove it and don't let ship shoot again
                ship.ball = None
            if not ob.passed and ship.x > ob.x:  # if player passes through obstacle, add another
                ob.passed = True
                add_obs = True
                ship.can_shoot = True  # gets its shot back once it passes through gate
                score += 1
            if ob.x + ob.OBS_TOP.get_width() < 0:  # if pipe leaves screen, add to array that will delete it
                obs_to_remove.append(ob)
            ob.move()
        if add_obs:  # add new obstacle, pin, and question here
            ob = Obstacle()
            obs.append(ob)
            pin = ob.pin
            pins.append(pin)
            question = random.choice(questions)
            question.shuffle_answers()
            question.reset_question(pin, buttons)
            update_game_speed(obs, pins, [top_base, bottom_base])
        for ob in obs_to_remove:
            obs.remove(ob)

        # Pin
        pins_to_remove = []
        for pin in pins:
            if pin.collide_with_ship(ship):
                run = False
            if ship.ball and pin.collide_with_ball(ship):
                if process_shot(question, ship.ball):
                    pins_to_remove.append(pin)
                ship.ball = None
            pin.move()
        for pin in pins_to_remove:
            pins.remove(pin)

        # Move base and background:
        if top_base.collide_with_ship(ship) or bottom_base.collide_with_ship(ship):
            run = False
        top_base.move()
        bottom_base.move()
        for bg in bgs:
            bg.move()

        draw_window(ship, top_base, bottom_base, bgs, obs, pins, ship.ball, buttons)

    game_over = True
    go_width = 600
    go_height = 500
    click = False
    title_font = pygame.font.SysFont('Segoe UI', 100, True)  # create font
    secondary_font = pygame.font.SysFont('sfprotextthin', 65)  # create font

    while game_over:
        pygame.draw.rect(win, colors['white'], (WINDOW_WIDTH/2 - (go_width+10)/2, WINDOW_HEIGHT / 2 - (go_height+10)/2,
                                                go_width+10, go_height+10))
        pygame.draw.rect(win, colors['menu dark purple'], (WINDOW_WIDTH/2-go_width/2, WINDOW_HEIGHT/2-go_height/2,
                                                           go_width, go_height))
        title_text = title_font.render('Game Over', True, colors['white'])  # create text
        win.blit(title_text, (WINDOW_WIDTH/2-title_text.get_width()/2, 180))  # draw title text

        mx, my = pygame.mouse.get_pos()  # get mouse position

        play_again_button = pygame.Rect(WINDOW_WIDTH/2 - BUTTON_WIDTH/2, 300, BUTTON_WIDTH, BUTTON_HEIGHT)
        play_again_button_color = colors['black']
        play_again_text = secondary_font.render('PLAY AGAIN', True, colors['white'])
        main_menu_button = pygame.Rect(WINDOW_WIDTH / 2 - BUTTON_WIDTH / 2, 400, BUTTON_WIDTH, BUTTON_HEIGHT)
        main_menu_button_color = colors['black']
        main_menu_text = secondary_font.render('MAIN MENU', True, colors['white'])
        exit_button = pygame.Rect(WINDOW_WIDTH / 2 - BUTTON_WIDTH / 2, 500, BUTTON_WIDTH, BUTTON_HEIGHT)
        exit_button_color = colors['black']
        exit_text = secondary_font.render('QUIT', True, colors['white'])
        if play_again_button.collidepoint((mx, my)):
            play_again_button_color = colors['dark gray']
            if click:  # restart game with same name
                active_window = 2
                break
        if main_menu_button.collidepoint((mx, my)):
            main_menu_button_color = colors['dark gray']
            if click:
                active_window = 0
                game_over = False
        if exit_button.collidepoint((mx, my)):
            exit_button_color = colors['dark gray']
            if click:  # if click on stop button
                active_window = -1
                game_over = False

        # Play Again button
        pygame.draw.rect(win, colors['white'],
                         (WINDOW_WIDTH/2 - (BUTTON_WIDTH + 10) / 2, 295, BUTTON_WIDTH + 10,
                          BUTTON_HEIGHT + 10))
        pygame.draw.rect(win, play_again_button_color, play_again_button)  # draw Play again button
        win.blit(play_again_text,
                 (WINDOW_WIDTH / 2 - play_again_text.get_width() / 2,
                  play_again_button.y + play_again_button.height / 2 - play_again_text.get_height() / 2))
        # Main menu button
        pygame.draw.rect(win, colors['white'],
                         (WINDOW_WIDTH / 2 - (BUTTON_WIDTH + 10) / 2, 395, BUTTON_WIDTH + 10, BUTTON_HEIGHT + 10))
        pygame.draw.rect(win, main_menu_button_color, main_menu_button)  # draw main_menu button
        win.blit(main_menu_text,
                 (WINDOW_WIDTH / 2 - main_menu_text.get_width() / 2,
                  main_menu_button.y + main_menu_button.height / 2 - main_menu_text.get_height() / 2))
        # Exit button
        pygame.draw.rect(win, colors['white'],
                         (WINDOW_WIDTH / 2 - (BUTTON_WIDTH + 10) / 2, 495, BUTTON_WIDTH + 10, BUTTON_HEIGHT + 10))
        pygame.draw.rect(win, exit_button_color, exit_button)  # draw exit button
        win.blit(exit_text,
                 (WINDOW_WIDTH / 2 - exit_text.get_width() / 2,
                  exit_button.y + exit_button.height / 2 - exit_text.get_height() / 2))
        click = False
        for event in pygame.event.get():
            if event.type == pygame.QUIT:  # if quit, don't start game and then quit
                active_window = -1
                game_over = False
            elif event.type == pygame.MOUSEBUTTONDOWN:  # if click down, set click to true
                if event.button == 1:
                    click = True

        pygame.display.update()

    with open('highscores.txt', 'a') as file:  # add score to highscores database
        file.write(f'{score}: {player_name}\n')
        file.close()


if __name__ == "__main__":
    while active_window != -1:
        if active_window == 0:
            active_window = start_screen()
        elif active_window == 1:
            leaderboard()
        elif active_window == 2:
            main()
    pygame.quit()
    pygame.font.quit()
    quit()
